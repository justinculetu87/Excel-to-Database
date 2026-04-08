import pandas as pd
import openpyxl
import numpy as np
import os

#load the master csv each time - if needed, uncomment the one below to add the interim data
df_main = pd.read_csv('P:\Housing Development Share\Development Cost Initiative\CDLAC_TCAC Competition Analysis\Application Analysis\Output\Interim Data\interim_analysis.csv')
df_main = df_main.dropna(how='all')  #removes rows with empty columns
#df_main = pd.read_csv('P:\Housing Development Share\Development Cost Initiative\CDLAC_TCAC Competition Analysis\Application Analysis\Output\Interim Data\Application_analysis.csv') #ONLY RUN IF REFORMATTING#

#add averages for later
if 'Averages' in df_main.iloc[:,0].values:
     df_main = df_main[df_main.iloc[:,0] != 'Averages']


##filter out non-unique funding sources


#functions to pull unique variables
def dollar_convert(value): #convert to integers to do calculations in cell transfer 
        if value is not None and isinstance (value, str) and '$' in value:
            return float(value.replace('$', '').replace (',', ''))
        return value or 0

def percent_convert(value): #convert to integers to do calculations in cell transfer 
        if value is not None and isinstance (value, str) and '%' in value:
            return float(value.replace('%', ''))
        return value or 0

def only_numeric(value):
    if value is None:
        return None
    try:
        float(str(value).replace('$', '').replace(',', '').replace('%', '').strip())
    except ValueError:
        return None
    
def safe_float(value):
    if value is None:
        return None
    try:
        return float(str(value).replace('$', '').replace(',', '').replace('%', '').strip())
    except (ValueError, TypeError):
        return None

#define workbooks
def application_workbooks(work_book, file_name):
    ws1 = work_book['Application'] 
    ws2 = work_book['Sources and Uses Budget']
    ws3 = work_book['Basis & Credits']
    ws4 = work_book['Sources and Basis Breakdown']
    ws5 = work_book['Tie Breaker']
    ws6 = work_book['CalHFA Addendum']

    #potential target values that are not always in the same cell
    deferred_target = None
    for row in ws1.iter_rows(min_row=627, max_row=638, min_col=3, max_col=41):
        for cell in row: 
            if cell.value and 'Developer Fee, Deferral' in str(cell.value):
                deferred_target = dollar_convert(ws1[f'AO{cell.row}'].value)
                break

    safehold_target = None
    for row in ws1.iter_rows(min_row=627, max_row=638, min_col=3, max_col=41):
        for cell in row:
            if cell.value and 'Safehold Inc.' in str(cell.value):
                safehold_target = ws1[f'AO{cell.row}'].value
                break

    haven_target = 0
    for row in ws1.iter_rows(min_row=627, max_row=638, min_col=3, max_col=41):
        for cell in row:
            if cell.value and 'haven' in str(cell.value):
                haven_target = dollar_convert(ws1[f'AO{cell.row}'].value)
                break

    ground_lease_target = None
    for row in ws1.iter_rows(min_row=627, max_row=638, min_col=3, max_col=41):
        for cell in row: 
            if cell.value and 'ground lease' in str(cell.value).lower():
                ground_lease_target = ws1[f'AO{cell.row}'].value
                break

    soft_fund_target = 0
    for row in ws1.iter_rows(min_row=627, max_row=638, min_col=3, max_col=41):
        for cell in row:
            if cell.value is not None and 'residual' in str(cell.value).lower():
                val_1 = dollar_convert(ws1[f'AO{cell.row}'].value)
                if val_1 is not None and isinstance(val_1,(int,float)):
                    soft_fund_target += val_1  #adds all of the soft funds
                break

    soft_fund_percent = 0
    for row in ws1.iter_rows(min_row=627, max_row=638, min_col=3, max_col=41):
        for cell in row: 
            if cell.value is not None and 'residual' in str(cell.value).lower():
                val_soft_percent = percent_convert(ws1[f'W{cell.row}'].value)
                if isinstance(val_soft_percent,(int,float)):
                    if val_soft_percent>soft_fund_percent: #takes the higher percentage
                        soft_fund_percent = val_soft_percent
                break

    perm_lender = None
    for row in ws1.iter_rows(min_row=627, max_row=638, min_col=3, max_col=41):
        for cell in row:
            if cell.value and 'required' in str(cell.value).lower():
                perm_lender = ws1[f'C{cell.row}'].value
                break

    perm_loan_target = 0
    for row in ws1.iter_rows(min_row=627, max_row=638, min_col=3, max_col=41):
        for cell in row:
            if cell.value is not None and 'required' in str(cell.value).lower():
                val_2 = dollar_convert(ws1[f'AO{cell.row}'].value)
                if val_2 is not None and isinstance(val_2,(int,float)):
                    perm_loan_target += val_2  #adds all of the permanent loans
                break 

    perm_loan_percent = None
    for row in ws1.iter_rows(min_row=627, max_row=638, min_col=3, max_col=41):
        for cell in row: 
            if cell.value is not None and 'required' in str(cell.value).lower():
                val_perm_percent = percent_convert(ws1[f'W{cell.row}'].value)
                if isinstance(val_perm_percent,(int,float)):
                    perm_loan_percent = val_perm_percent 

    parking_spaces = safe_float(ws1['Q494'].value) if safe_float(ws1['Q494'].value) else (safe_float(ws1['M495'].value) or 0) + (safe_float(ws1['AH495'].value) or 0)
    

    new_row = { #call the cell values
        'Application #': file_name, 
        'Applicant Name': ws1['H16'].value if ws1['H16'].value is not None else '',
        'Project Name': ws1['H18'].value if ws1['H18'].value is not None else '', 
        'Project Type': ws1['D211']. value if ws1['D211'].value is not None else '',
        'Requested Credit': ws1['D355'].value if ws1['M355'].value not in (None, 'N/A', 'NA') else ws1['D356'].value if ws1['M356'].value not in (None, 'N/A', 'NA') else ws1['D357'].value 
        if ws1['M357'] not in (None, 'N/A', 'NA') else ws1['D358'].value if ws1['M358'] not in (None, 'N/A', 'NA') else'',
        'Geographic':  ws1['D220'].value if ws1['D220'].value is not None else '',
        'Total Units': ws1['AG437'].value if ws1['AG437'].value is not None else '',
        'Acreage': f"{ws1['P418'].value:.1f}" if ws1['P418'].value is not None else '',
        'Density (DU/AC)': f"{ws1['AF418'].value:.2f}" if ws1['AF418'].value is not None else '',
        'Number of Stories': ws1['N372'].value if ws1['N372'].value not in (None, 'N/A', 'NA') else ws1['AD411'].value if ws1['AD411'].value not in (None, 'N/A', 'NA') else ws1['AD412'].value,  
        'Net Rentable': ws1['AG442'].value if ws1['AG442'].value is not None else '',
        'Parking Structure': ws1['AG449'].value if ws1['AG449'].value is not None else '',
        'Total Gross Sq Ft': ws1['AG450'].value if ws1['AG450'].value is not None else '',
        'Land Cost – Total ($)': f"${dollar_convert(ws1['Q391'].value)}",
        'Name of Land Seller': ws1['J385'].value if ws1['J385'].value is not None else '',
        'Land Seller Address': ws1['J388'].value if ws1['J388'].value is not None else '',
        'Land Seller Phone #': ws1['I392'].value if ws1['I392'].value is not None else '',
        'Signatory of Seller': ws1['AC385'].value if ws1['AC385'].value is not None else '',
        'Land Cost / Unit ($)': f"${dollar_convert(ws1['Q391'].value) / ws1['AG437'].value:.0f}" if ws1['Q391'].value not in (None, 0, 'N/A') and ws1['AG437'].value not in (None, 0, 'N/A') else '', 
        'Land Cost / Acre ($)': f"${dollar_convert(ws1['Q391'].value) / dollar_convert(ws1['P418'].value):.0f}" if ws1['Q391'].value not in (None, 0, 'N/A') and ws1['P418'].value not in (None, 0, 'N/A') else '',
        'Ground Lease Proceeds': f"${dollar_convert(ground_lease_target or haven_target or safehold_target):.0f}" if (ground_lease_target or haven_target or safehold_target) not in (None, 0) else '',  
        'Parking Spaces': safe_float(ws1['Q494'].value) if safe_float(ws1['Q494'].value) and safe_float(ws1['Q494'].value) not in (None, 0, 'N/A') else (safe_float(ws1['M495'].value) or 0) + (safe_float(ws1['AH495'].value) or 0), 
        'Parking Ratio': f"{(parking_spaces / ws1['AG439'].value) * 100:.2f}%" if ws1['AG439'].value not in (None, 0, 'N/A') else '', 
        #'Parking Type': enter after running , 
        '1BR Units (%)': f"{ws1['AB987'].value} ({ws1['AB987'].value / ws1['AG437'].value * 100:.0f}%)" if ws1['AG437'].value is not None else '',
        '2BR Units (%)': f"{ws1['AB988'].value} ({ws1['AB988'].value / ws1['AG437'].value * 100:.0f}%)" if ws1['AG437'].value is not None else '', 
        '3BR Units (%)': f"{ws1['AB989'].value} ({ws1['AB989'].value / ws1['AG437'].value * 100:.0f}%)" if ws1['AG437'].value is not None else '', 
        '4BR Units (%)': f"{ws1['AB990'].value} ({ws1['AB990'].value / ws1['AG437'].value * 100:.0f}%)" if ws1['AG437'].value is not None else '', 
        'AVG AMI': f"{ws1['AC753'].value * 100:.2f}%", 
        'Hard Costs incl. Contingency ($)': f"${(dollar_convert(ws2['B38'].value) + dollar_convert(ws2['B26'].value)) + dollar_convert(ws2['B79'].value)}",   
        'Hard Cost / Unit ($)': f"${((dollar_convert(ws2['B38'].value) + dollar_convert(ws2['B26'].value)) + dollar_convert(ws2['B79'].value)) / ws1['AG437'].value:.0f}" if ws1['AG437'].value not in (None, 0) else '' , 
        'Hard Cost Net Rentable': f"${((dollar_convert(ws2['B38'].value) + dollar_convert(ws2['B26'].value)) + dollar_convert(ws2['B79'].value)) / ws1['AG442'].value:.1f}" if ws1['AG442'].value not in (None, 0) else '', 
        'Hard Cost Gross SF': f"${((dollar_convert(ws2['B38'].value) + dollar_convert(ws2['B26'].value)) + dollar_convert(ws2['B79'].value)) / ws1['AG450'].value:.0f}" if ws1['AG450'].value not in (None, 0) else '' , 
        'Prevailing Wage': ws5['G120'].value if ws5['G120'].value is not None else '',
        'OPX / Unit ($)': f"${dollar_convert(ws1['AC885'].value)}", 
        'Soft Funding ($)': f"${dollar_convert(soft_fund_target):.0f}" if soft_fund_target not in (None, 'N/A') else '',  #add all of the soft funding
        'Soft Funding Rate': f"{soft_fund_percent * 100:.2f}%" if soft_fund_percent is not None else '', #take top % in residual
        'Soft Funding / Unit ($)': f"${dollar_convert(soft_fund_target) / ws1['AG437'].value:.0f}" if ws1['AG437'].value not in (None, 0) else '', 
        'Construction Interest': f"${dollar_convert(ws2['B45'].value)}" if ws2['B45'].value is not None else '',   
        #'B- Bond ': add after running , 
        #'GP Note': add after running ,
        'Total Developer Cost': f"${dollar_convert(ws2['B104'].value)}" if ws2['B104'].value is not None else '',  
        'Deferred Developer Fee': f"${dollar_convert(deferred_target):.0f}" if deferred_target not in (None, 'N/A') else '', 
        'Cash Fee': f"${dollar_convert(ws2['B104'].value) - deferred_target:.0f}" if deferred_target not in (None, 'N/A') else '', #(Total Dev Cost - Deferred Developer Fee)
        'Total Development Cost ($)': f"${dollar_convert(ws3['AB47'].value)}" if ws3['AB47'].value is not None else '', 
        'TDC / Unit ': f"${dollar_convert(ws3['AB47'].value) / ws1['AG437'].value:.0f}" if ws1['AG437'].value not in (None, 0) else '',
        'State Tax Credits': f"${dollar_convert(ws3['AB79'].value)}" if ws3['AB79'].value is not None else ' ',
        'State Tax Credit Factor': f"${dollar_convert(ws3['AB72'].value)}" if ws3['AB72'].value is not None else ' ', 
        'Perm Lender': perm_lender if perm_lender is not None else '',  
        'Perm Loan Amount': f"${dollar_convert(perm_loan_target):.0f}" if perm_loan_target is not None else '',  
        'Rate': f"{perm_loan_percent * 100:.2f}%" if perm_loan_percent not in (None, 0) else '', 
        'Equity from Federal Credit': f"${dollar_convert(ws3['AB57'].value)}" if ws3['AB57'].value is not None else '', 
        'Federal Tax Credit Factor': f"${dollar_convert(ws3['AB50'].value):.3f}" if ws3['AB50'].value is not None else '' , 
        'Equity Investor': ws1['AA311'].value if ws1['AA311'].value is not None else '', 
        'Architect': ws1['AA287'].value if ws1['AA287'].value is not None else '' , 
        'General Contractor': ws1['AA295'].value if ws1['AA295'].value is not None else '', 
        'Resource Area': ws1['T193'].value if ws1['T193'].value is not None else '', 
        'CUAC Analyst': ws1['AA303'].value if ws1['AA303'] is not None else '', 
        'Tie Breaker': f"{ws5['H5'].value * 100 :.2f}%"
    }
    return new_row

#add new files
add_application = input('Load in new applications (yes/no): ')

new_obs = []

if add_application.lower() == 'yes':
    add_method = input('Single file entry or multiple file entry? (single/multiple): ')  #which method is being used
    
    if add_method.lower() == 'single':  #for single method
         while True:
            file_path = input('Enter the path to the application file: ')
            try:  #in case file path is entered incorrectly
                work_book = openpyxl.load_workbook(file_path, data_only=True)
            except FileNotFoundError:
                print('File not found. Enter new path.')
                continue
            except Exception as e:
                print(f'Error opening file: {e}. Re-enter file path.')
                continue

            new_row = application_workbooks(work_book, os.path.basename(file_path)) #add the new application to the excel
            new_obs.append(new_row)
           

            #adding another single observation
            add_another = input('Add another application? (yes/no): ')  #adding another single file?
            if add_another.lower() !='yes':
                break

    elif add_method.lower() == 'multiple':   #entry of multiple files at once
        base_paths = []  #enter path to the Round of applications
        while True:
            enter_folder = input('Enter a folder path (press Enter to stop): ') #'Enter' starts the next process
            if enter_folder == '':
                break
            base_paths.append(enter_folder)

        for base_path in base_paths:  #find the files that should be downloaded
            files = [f for f in os.listdir(base_path) if (f.startswith('25') or f.startswith('24')) and (f.endswith('xlsx') or f.endswith('xlsm'))]
            for file in files:
                file_path = os.path.join(base_path, file)
                try:
                    work_book = openpyxl.load_workbook(file_path, data_only=True)
                except Exception as e:
                    print(f'Error opening {file}: {e}. Skip')
                    continue

                new_row = application_workbooks(work_book, file)
                new_obs.append(new_row)

#add to dataframe
if new_obs:
    df_main = pd.concat([df_main, pd.DataFrame(new_obs)], ignore_index=True)

#########################################################################################################

#get averages
average = {}
for col in df_main.columns:
    try:
        strings = df_main[col].astype(str) #cast all to strings

        #determine formats
        has_dollar = strings.str.contains(r'\$', regex=True).any()
        has_percent = strings.str.contains(r'\%', regex=True).any()
        percent_parenthesis = strings.str.contains(r'\(\d+%\)', regex=True).any()

        if percent_parenthesis: #only pull numbers no matter the formatting of the cell 
            cleaned = pd.to_numeric(strings.str.extract(r'\((\d+)%\)')[0], errors='coerce')
        else:
            cleaned = pd.to_numeric(strings.str.replace(r'[$,%]', '', regex=True), errors='coerce')


        if cleaned.notna().any():  #get the averages 
            mean_val = cleaned.mean()
            if has_dollar:
                average[col] = f'${mean_val:,.1f}'
            elif has_percent:
                average[col] = f'{mean_val:,.1f}%'
            elif percent_parenthesis:
                average[col] = f'({mean_val:,.0f}%)'
            else: 
                average[col] = round(mean_val,1)
        else:
            average[col] = ''
    except Exception:
        average[col] = np.nan #empty cell if not applicable

df_main = pd.concat([df_main, pd.DataFrame([average])], ignore_index = True)
df_main.iloc[-1, 0] = 'Averages'

#Ask if user wants to save to winning or losing master file
master_type = input('Save to winning or losing analysis (winning/losing): ' )
if master_type.lower() == 'winning':
    df_main.to_excel('P:\Housing Development Share\Development Cost Initiative\CDLAC_TCAC Competition Analysis\Application Analysis\Output\Master File\winning_analysis_master_25.xlsx', index=False)
    print('File saved successfully')

if master_type.lower() == "losing":
    df_main.to_excel('P:\Housing Development Share\Development Cost Initiative\CDLAC_TCAC Competition Analysis\Application Analysis\Output\Master File\losing_analysis_master_25.xlsx', index=False)
    print('File saved successfully')


    






